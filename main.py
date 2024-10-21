import numpy as np
import pandas as pd
import csv
import time
import math

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from typing import List
from typing import Tuple

from stage import Engine, STS
print("\n")
payload_mass = 1500
dv_min = 12512
engines = None
# Find the number of engines in each stage assuming only n engines are allowed.
# stolen from https://stackoverflow.com/questions/13131491/partition-n-items-into-k-bins-in-python-lazily
def partition(lista,bins):
    if len(lista)==1 or bins==1:
        yield [lista]
    elif len(lista)>1 and bins>1:
        for i in range(1,len(lista)):
            for part in partition(lista[i:],bins-1):
                if len([lista[:i]]+part)==bins:
                    yield [lista[:i]]+part
    
engine_count_combinations = partition(range(1,7), 4)
engine_count_combinations = [[1] + [len(i) for i in combination] for combination in engine_count_combinations]
engine_count_combinations = [tuple([count * multiplier for multiplier, count in zip([1, 1, 2, 2, 4], combination)]) for combination in engine_count_combinations]
max_engines_per_stage = [1, 1, 4, 8, 16]
engine_count_combinations = list(filter(lambda count_combination: \
                                   all(engine_count <= max_engine_count for engine_count, max_engine_count in zip(count_combination, max_engines_per_stage)) \
                                    , engine_count_combinations))
print(engine_count_combinations)
print(len(engine_count_combinations))

exit()

# getting the engines
with open("engines.csv") as engine_data:
    reader = csv.reader(engine_data)
    reader.__next__()
    reader.__next__()
    engines: List[Engine] = [Engine.from_row(row) for row in reader]

engines_per_stage: List[List[Engine]] = []

engines_per_stage.append([Engine.payload(payload_mass)])
engines_per_stage.append(list(filter(lambda e: "LOX" in e.group or "NTO" in e.group or "Hybrid" in e.group or e.group == "EP" or e.group == "None", engines)))
engines_per_stage.append(list(filter(lambda e: "LOX" in e.group or "NTO" in e.group or "Hybrid" in e.group, engines)))
engines_per_stage.append(list(filter(lambda e: e.group == "Solid" or "LOX" in e.group or "NTO" in e.group or "Hybrid" in e.group, engines)))
# engines_per_stage.append(list(filter(lambda e: e.group == "Solid", engines)))
engines_per_stage.append(list(filter(lambda e: e.group == "Solid" or e.group == "None", engines)))

staging_infos = engines_per_stage
staging_infos.append(engine_count_combinations)

stage_options = pd.MultiIndex.from_product(staging_infos, names=["PL", "S5", "S4", "S3", "S2", "S1", "conf"])
print([len(l) for l in engines_per_stage])

min_cost_sts: List[Tuple[float, STS]] = []
min_cost = 1000000000
max_dv_sts = None
max_dv = 0

i = 0
i_bad = 0
start_time = time.process_time()
for staging_info in stage_options:
    engines_initial = staging_info[0:-1]
    engine_counts = staging_info[-1]
    sts = STS(engines_initial, engine_counts)
    if any([e.group == "None" and n > 1 for e,n in zip(engines_initial, engine_counts)]):
        i_bad += 1
        continue
    if(math.isnan(sts.delta_V())):
        i_bad += 1
        continue
    dv = sts.delta_V()
    total_mass = sts.payload_mass / sts.mass_fraction()
    # The cost function that drives which setups are optimal
    cost = total_mass * (1 + 0.5 * sts.engine_count())
    if cost < min_cost and dv > dv_min and all([abs(cost - sts_cost) > 5000 for sts_cost, _ in min_cost_sts]):
        min_cost_sts.append((cost, sts))
        min_cost_sts = sorted(min_cost_sts, key = lambda p: -p[0])
        if len(min_cost_sts) > 10:
            min_cost_sts.pop()
            min_cost = max(cost for cost, _ in min_cost_sts)
    if dv > max_dv:
        max_dv_sts = sts
        max_dv = dv
    i += 1
    if i % 10000 == 0:
        print(f"\n#{i} + {i_bad}:\nmin cost = {min_cost}, with \n\t{[sts for _, sts in min_cost_sts]}\ndv_max = {max_dv} with \n\t{max_dv_sts}")
end_time = time.process_time()

def describe_stages(wb: Workbook, sheet_name, index: int, sts: STS):
    if sts is not None:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name, -1)
        print("\n")
        print(sts)
        engine_data
        names = [e.name for e in sts.engines] + ["Total"]
        n = list(sts.engine_counts) + [sts.engine_count()]
        dv = [s.delta_v for s in sts.stages] + [sts.delta_V()]
        m = [e.mass_gross for e in sts.engines] + [sts.payload_mass / sts.mass_fraction()]
        mu = [s.mass_fraction for s in sts.stages] + [np.nan]
        Isp = [e.specific_impulse for e in sts.engines] + [np.nan]
        df = pd.DataFrame(zip(names, n, dv, m, mu, Isp), columns=["Name", "Engine Count", "Delta V", "Gross Mass", "Mu", "Specific Impulse"])
        for row in dataframe_to_rows(df, index=True, header=True):
            sheet.append(row)
        print(df)

wb = Workbook()

print("\nMinimum Cost:")
print(min_cost) 
[describe_stages(wb, "Min Cost", i, sts) for i, (cost, sts) in enumerate(min_cost_sts)] 
    
print("\nMax dV:")
describe_stages(wb,"Max DV", 0, max_dv_sts)

wb.save("output.xlsx")

dt = end_time - start_time

print(f"Calculated {[len(l) for l in staging_infos]} = total {stage_options.size} cases = {i} valid and {i_bad} invalid in {dt} seconds = {int(stage_options.size / dt)} cases/second")
